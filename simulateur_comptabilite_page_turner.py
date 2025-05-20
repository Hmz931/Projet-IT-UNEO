import pandas as pd
import numpy as np
import os
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import locale
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Set locale for French-Swiss date formats
try:
    locale.setlocale(locale.LC_TIME, 'fr_CH.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'fr_CH')
    except:
        logging.warning("Failed to set French-Swiss locale, using default.")

# Simulation parameters
class SimulationParams:
    def __init__(self, start_year=2021, end_year=2025):
        self.start_year = start_year
        self.end_year = end_year
        self.tva_rates = {2021: 7.7, 2022: 7.7, 2023: 7.7, 2024: 8.1, 2025: 8.1}

# Data generator
class AccountingDataGenerator:
    def __init__(self, params):
        self.params = params
        self.init_base_data()

    def init_base_data(self):
        """Initialize base data for a bookstore"""
        # Chart of accounts
        self.accounts = {
            1000: {"name": "Caisse", "category": "Actif", "sous_classe": "Actif circulant", "detail_categorie": "Liquidité", "type_compte": "Centralisateur"},
            1010: {"name": "Compte Banque", "category": "Actif", "sous_classe": "Actif circulant", "detail_categorie": "Banque", "type_compte": "Centralisateur"},
            1100: {"name": "Créances clients", "category": "Actif", "sous_classe": "Actif circulant", "detail_categorie": "Créances", "type_compte": "Centralisateur"},
            1170: {"name": "TVA à récupérer", "category": "Actif", "sous_classe": "Actif circulant", "detail_categorie": "TVA", "type_compte": "TVA"},
            1171: {"name": "TVA ajustements", "category": "Actif", "sous_classe": "Actif circulant", "detail_categorie": "TVA", "type_compte": "TVA"},
            1172: {"name": "TVA taux réduit", "category": "Actif", "sous_classe": "Actif circulant", "detail_categorie": "TVA", "type_compte": "TVA"},
            1200: {"name": "Stocks de marchandises", "category": "Actif", "sous_classe": "Actif circulant", "detail_categorie": "Stocks", "type_compte": "Centralisateur"},
            1500: {"name": "Équipements magasin", "category": "Actif", "sous_classe": "Actif immobilisé", "detail_categorie": "Immobilisations", "type_compte": "Centralisateur"},
            2000: {"name": "Dettes fournisseurs", "category": "Passif", "sous_classe": "Dettes à court terme", "detail_categorie": "Dettes", "type_compte": "Centralisateur"},
            2200: {"name": "TVA due", "category": "Passif", "sous_classe": "Autres dettes à court terme", "detail_categorie": "TVA", "type_compte": "TVA"},
            2270: {"name": "AVS/AI/APG", "category": "Passif", "sous_classe": "Autres dettes à court terme", "detail_categorie": "Charges sociales", "type_compte": "Centralisateur"},
            2271: {"name": "LAA", "category": "Passif", "sous_classe": "Autres dettes à court terme", "detail_categorie": "Charges sociales", "type_compte": "Centralisateur"},
            2272: {"name": "IJM", "category": "Passif", "sous_classe": "Autres dettes à court terme", "detail_categorie": "Charges sociales", "type_compte": "Centralisateur"},
            2273: {"name": "LPP", "category": "Passif", "sous_classe": "Autres dettes à court terme", "detail_categorie": "Charges sociales", "type_compte": "Centralisateur"},
            2279: {"name": "Impôt à la source", "category": "Passif", "sous_classe": "Autres dettes à court terme", "detail_categorie": "Charges sociales", "type_compte": "Centralisateur"},
            2299: {"name": "Salaires à payer", "category": "Passif", "sous_classe": "Autres dettes à court terme", "detail_categorie": "Technique", "type_compte": "Centralisateur"},
            2800: {"name": "Capital social", "category": "Passif", "sous_classe": "Fonds propres", "detail_categorie": "Capital", "type_compte": "Centralisateur"},
            2979: {"name": "Bénéfice/perte", "category": "Passif", "sous_classe": "Réserves / bénéfices et pertes", "detail_categorie": "Résultat", "type_compte": "Centralisateur"},
            3400: {"name": "Ventes de prestations", "category": "Produit", "sous_classe": "Chiffre d'affaire", "detail_categorie": "Ventes", "type_compte": "Centralisateur"},
            3410: {"name": "Ventes en ligne", "category": "Produit", "sous_classe": "Chiffre d'affaire", "detail_categorie": "Ventes", "type_compte": "Centralisateur"},
            3600: {"name": "Autres ventes", "category": "Produit", "sous_classe": "Chiffre d'affaire", "detail_categorie": "Ventes", "type_compte": "Centralisateur"},
            4000: {"name": "Achats livres", "category": "Charge", "sous_classe": "Charges de matériel", "detail_categorie": "Marchandises", "type_compte": "Centralisateur"},
            4201: {"name": "Frais de transport", "category": "Charge", "sous_classe": "Charges de matériel", "detail_categorie": "Marchandises", "type_compte": "Centralisateur"},
            5200: {"name": "Salaires de base", "category": "Charge", "sous_classe": "Charges salariales", "detail_categorie": "Salaires", "type_compte": "Centralisateur"},
            5270: {"name": "AVS, AI, APG", "category": "Charge", "sous_classe": "Charges sociales", "detail_categorie": "Charges sociales", "type_compte": "Centralisateur"},
            5283: {"name": "Frais de repas", "category": "Charge", "sous_classe": "Autres charges de personnel", "detail_categorie": "Personnel", "type_compte": "Centralisateur"},
            6000: {"name": "Loyers", "category": "Charge", "sous_classe": "Charges de locaux", "detail_categorie": "Locaux", "type_compte": "Centralisateur"},
            6040: {"name": "Nettoyage", "category": "Charge", "sous_classe": "Charges de locaux", "detail_categorie": "Locaux", "type_compte": "Centralisateur"},
            6500: {"name": "Frais administratifs", "category": "Charge", "sous_classe": "Charges d'administration", "detail_categorie": "Administration", "type_compte": "Centralisateur"},
            6510: {"name": "Téléphone et internet", "category": "Charge", "sous_classe": "Charges d'administration", "detail_categorie": "Administration", "type_compte": "Centralisateur"},
            6600: {"name": "Publicité", "category": "Charge", "sous_classe": "Charges de publicité", "detail_categorie": "Publicité", "type_compte": "Centralisateur"},
            6643: {"name": "Cadeaux clients", "category": "Charge", "sous_classe": "Charges de publicité", "detail_categorie": "Publicité", "type_compte": "Centralisateur"},
            6700: {"name": "Autres charges", "category": "Charge", "sous_classe": "Autres charges d'exploitation", "detail_categorie": "Exploitation", "type_compte": "Centralisateur"},
            6800: {"name": "Amortissements", "category": "Charge", "sous_classe": "Amortissements", "detail_categorie": "Amortissements", "type_compte": "Centralisateur"},
            6900: {"name": "Charges financières", "category": "Charge", "sous_classe": "Charges et produits financiers", "detail_categorie": "Financier", "type_compte": "Centralisateur"},
            8200: {"name": "Charges exceptionnelles", "category": "Charge", "sous_classe": "Résultats extraordinaires", "detail_categorie": "Exceptionnel", "type_compte": "Centralisateur"},
            8510: {"name": "Produits exceptionnels", "category": "Produit", "sous_classe": "Résultats extraordinaires", "detail_categorie": "Exceptionnel", "type_compte": "Centralisateur"},
            8900: {"name": "Impôts directs", "category": "Charge", "sous_classe": "Clôture", "detail_categorie": "Impôts", "type_compte": "Centralisateur"}
        }
        # Analytical codes
        self.analytical_codes = {
            "A001": {"libelle": "Ventes magasin", "type": "Produit"},
            "A002": {"libelle": "Ventes en ligne", "type": "Produit"},
            "A003": {"libelle": "Événements", "type": "Produit"}
        }
        # Suppliers
        self.suppliers = {
            1: {"nom": "Payot Librairie", "adresse": "Rue de la Confédération 7", "code_postal": "1204", "pays": "CH", "devise_facture": "CHF"},
            2: {"nom": "Libra Diffusion", "adresse": "Avenue de France 12", "code_postal": "1004", "pays": "CH", "devise_facture": "CHF"}
        }
        # Clients
        self.clients = {
            1: {"nom": "École de Genève", "adresse": "Rue des Écoles 10", "code_postal": "1205", "pays": "CH"},
            2: {"nom": "Club de Lecture SA", "adresse": "Avenue de la Paix 5", "code_postal": "1202", "pays": "CH"}
        }
        # Currencies
        self.currencies = {
            "CHF": {"nom": "Franc Suisse"},
            "EUR": {"nom": "Euro"},
            "USD": {"nom": "Dollar Américain"}
        }

    def get_random_date(self, year, month=None):
        """Generate a random date in the given year and month"""
        if month:
            start_date = datetime(year, month, 1)
            end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)
        else:
            start_date = datetime(year, 1, 1)
            end_date = datetime(year, 12, 31)
        days_range = (end_date - start_date).days
        random_day = random.randint(0, days_range)
        return start_date + timedelta(days=random_day)

    def generate_journal_entry(self, date, compte_debit, compte_credit, montant_debit, montant_credit, libelle, code_analytique="", ref_document=""):
        """Generate a journal entry"""
        if not (compte_debit in self.accounts and compte_credit in self.accounts):
            logging.error(f"Invalid account: Debit {compte_debit}, Credit {compte_credit}")
            return None
        return {
            "Date": date.strftime("%d.%m.%Y"),
            "CompteDebit": compte_debit,
            "CompteCredit": compte_credit,
            "MontantDebit": round(montant_debit, 2) if montant_debit else np.nan,
            "MontantCredit": round(montant_credit, 2) if montant_credit else np.nan,
            "Libelle": libelle,
            "CodeAnalytique": code_analytique if code_analytique in self.analytical_codes else "",
            "RefDocument": ref_document
        }

    def generate_supplier_invoice(self, year, numero_document):
        """Generate a supplier invoice"""
        supplier_id = random.randint(1, len(self.suppliers))
        supplier = self.suppliers[supplier_id]
        date_facture = self.get_random_date(year)
        date_paiement = date_facture + timedelta(days=random.randint(10, 30))
        montant = round(random.uniform(500, 5000), 2)
        has_vat = True
        vat_rate = self.params.tva_rates[year]
        amount_vat = round(montant * vat_rate / 100, 2)
        amount_ht = montant
        statut_facture = "ERLED" if random.random() < 0.9 else "OFFEN"
        type_document = "F"
        numero_facture = f"F-{year}-{numero_document:04d}"
        ref_document = f"YOOZ{year}{numero_document:04d}"
        vat_account = random.choice([1170, 1171, 1172])
        journal_entries = [
            self.generate_journal_entry(date_facture, 4000, 2000, amount_ht, np.nan, f"Facture {numero_facture} - {supplier['nom']}", "", ref_document),
            self.generate_journal_entry(date_facture, vat_account, 2000, amount_vat, np.nan, f"TVA sur facture {numero_facture}", "", ref_document)
        ]
        if statut_facture == "ERLED":
            journal_entries.append(self.generate_journal_entry(date_paiement, 2000, 1010, np.nan, amount_ht + amount_vat, f"Paiement facture {numero_facture}", "", ref_document))
        # Occasional transport fee (4201, ~20% of invoices)
        if random.random() < 0.2:
            transport_fee = round(random.uniform(50, 200), 2)
            journal_entries.append(self.generate_journal_entry(
                date_facture, 4201, 2000, transport_fee, np.nan,
                f"Frais de transport facture {numero_facture}",
                "", ref_document
            ))
            if statut_facture == "ERLED":
                journal_entries.append(self.generate_journal_entry(
                    date_paiement, 2000, 1010, np.nan, transport_fee,
                    f"Paiement frais de transport {numero_facture}",
                    "", ref_document
                ))
        return {
            "NumeroDocument": numero_document,
            "IDFournisseur": supplier_id,
            "DateFacture": date_facture.strftime("%d.%m.%Y"),
            "NumeroFacture": numero_facture,
            "DatePaiement": date_paiement.strftime("%d.%m.%Y") if statut_facture == "ERLED" else None,
            "Montant": montant + amount_vat,
            "Monnaie": supplier['devise_facture'],
            "StatutFacture": statut_facture,
            "TypeDocument": type_document,
            "RefDocument": ref_document
        }, journal_entries

    def generate_client_invoice(self, year, numero_document):
        """Generate a client invoice"""
        client_id = random.randint(1, len(self.clients))
        client = self.clients[client_id]
        date_facture = self.get_random_date(year)
        date_paiement = date_facture + timedelta(days=random.randint(5, 20))
        montant = round(random.uniform(200, 2000), 2) if year in [2021, 2022] else round(random.uniform(500, 5000), 2)
        has_vat = True
        vat_rate = self.params.tva_rates[year]
        amount_vat = round(montant * vat_rate / (100 + vat_rate), 2)
        amount_ht = montant - amount_vat
        statut_facture = "ERLED" if random.random() < 0.85 else "OFFEN"
        type_document = "F"
        numero_facture = f"C-{year}-{numero_document:04d}"
        ref_document = f"CLI{year}{numero_document:04d}"
        sales_account = 3400 if random.random() < 0.7 else 3410
        code_analytique = "A001" if sales_account == 3400 else "A002"
        journal_entries = [
            self.generate_journal_entry(date_facture, 1100, sales_account, amount_ht, np.nan, f"Facture {numero_facture} - {client['nom']}", code_analytique, ref_document),
            self.generate_journal_entry(date_facture, 1100, 2200, amount_vat, np.nan, f"TVA sur facture {numero_facture}", code_analytique, ref_document)
        ]
        if statut_facture == "ERLED":
            journal_entries.append(self.generate_journal_entry(date_paiement, 1010, 1100, np.nan, montant, f"Encaissement facture {numero_facture}", code_analytique, ref_document))
        return {
            "NumeroDocument": numero_document,
            "IDClient": client_id,
            "DateFacture": date_facture.strftime("%d.%m.%Y"),
            "NumeroFacture": numero_facture,
            "DatePaiement": date_paiement.strftime("%d.%m.%Y") if statut_facture == "ERLED" else None,
            "Montant": montant,
            "Monnaie": "CHF",
            "StatutFacture": statut_facture,
            "TypeDocument": type_document,
            "RefDocument": ref_document
        }, journal_entries

    def generate_salary_entries(self, year):
        """Generate monthly salary entries with specified accounts"""
        journal_entries = []
        for month in range(1, 13):
            date = datetime(year, month, random.randint(25, 28))
            gross_salary = round(random.uniform(3000, 5000), 2) * 5  # 5 employees
            # Social charges breakdown (total ~15% of gross salary)
            avs_ac_amat = round(gross_salary * 0.06, 2)  # AVS/AC/AMAT (6%)
            laa = round(gross_salary * 0.02, 2)          # LAA (2%)
            ijm = round(gross_salary * 0.01, 2)          # IJM (1%)
            lpp = round(gross_salary * 0.03, 2)          # LPP (3%)
            impot_source = round(gross_salary * 0.03, 2) # IS (3%)
            total_social_charges = avs_ac_amat + laa + ijm + lpp + impot_source
            net_salary = gross_salary + total_social_charges
            ref_document = f"SAL-{year}-{month:02d}"
            # Salary postings
            journal_entries.append(self.generate_journal_entry(
                date, 5200, 2270, gross_salary * 0.06, np.nan,
                f"Salaire - AVS/AC/AMAT mois {month}",
                "", ref_document
            ))
            journal_entries.append(self.generate_journal_entry(
                date, 5200, 2271, laa, np.nan,
                f"Salaire - LAA mois {month}",
                "", ref_document
            ))
            journal_entries.append(self.generate_journal_entry(
                date, 5200, 2272, ijm, np.nan,
                f"Salaire - IJM mois {month}",
                "", ref_document
            ))
            journal_entries.append(self.generate_journal_entry(
                date, 5200, 2273, lpp, np.nan,
                f"Salaire - LPP mois {month}",
                "", ref_document
            ))
            journal_entries.append(self.generate_journal_entry(
                date, 5200, 2279, impot_source, np.nan,
                f"Salaire - IS mois {month}",
                "", ref_document
            ))
            journal_entries.append(self.generate_journal_entry(
                date, 5200, 2299, gross_salary, np.nan,
                f"Salaire - Salaire à payer mois {month}",
                "", ref_document
            ))
            # Payment (clears 2299)
            journal_entries.append(self.generate_journal_entry(
                date, 2299, 1010, np.nan, net_salary,
                f"Paiement salaire mois {month}",
                "", ref_document
            ))
            # Occasional expense (5283, ~30% of months)
            if random.random() < 0.3:
                expense = round(random.uniform(50, 150), 2)
                journal_entries.append(self.generate_journal_entry(
                    date, 5283, 1010, expense, np.nan,
                    f"Frais de repas mois {month}",
                    "", ref_document
                ))
        return journal_entries

    def generate_vat_settlement(self, journal_df, year):
        """Generate quarterly VAT settlements"""
        journal_entries = []
        vat_accounts = [1170, 1171, 1172, 2200]
        for quarter in range(1, 5):
            start_date = datetime(year, (quarter - 1) * 3 + 1, 1)
            end_date = (datetime(year, (quarter - 1) * 3 + 3, 1) + timedelta(days=31)).replace(day=1) - timedelta(days=1)
            quarter_entries = journal_df[
                (journal_df["Date"].apply(lambda x: datetime.strptime(x, "%d.%m.%Y")) >= start_date) &
                (journal_df["Date"].apply(lambda x: datetime.strptime(x, "%d.%m.%Y")) <= end_date)
            ]
            vat_balances = {}
            for vat_account in vat_accounts:
                debit_sum = quarter_entries[quarter_entries["CompteDebit"] == vat_account]["MontantDebit"].sum()
                credit_sum = quarter_entries[quarter_entries["CompteCredit"] == vat_account]["MontantCredit"].sum()
                vat_balances[vat_account] = debit_sum - credit_sum
            net_vat = vat_balances[2200] - sum(vat_balances[acc] for acc in [1170, 1171, 1172])
            settlement_date = end_date + timedelta(days=30)
            ref_document = f"TVA-{year}-Q{quarter}"
            # Pay or receive VAT
            if net_vat > 0:
                journal_entries.append(self.generate_journal_entry(
                    settlement_date, 2200, 1010, np.nan, net_vat,
                    f"Paiement TVA Q{quarter} {year}",
                    "", ref_document
                ))
            elif net_vat < 0:
                journal_entries.append(self.generate_journal_entry(
                    settlement_date, 1010, 1170, abs(net_vat), np.nan,
                    f"Remboursement TVA Q{quarter} {year}",
                    "", ref_document
                ))
            # Clear VAT accounts
            for vat_account in vat_accounts:
                balance = vat_balances[vat_account]
                if vat_account == 2200 and net_vat > 0:
                    continue  # Already cleared by payment
                elif vat_account == 1170 and net_vat < 0:
                    continue  # Already adjusted by refund
                if balance != 0:
                    if balance > 0:
                        journal_entries.append(self.generate_journal_entry(
                            settlement_date, vat_account, 1010, np.nan, abs(balance),
                            f"Régularisation TVA compte {vat_account} Q{quarter}",
                            "", ref_document
                        ))
                    else:
                        journal_entries.append(self.generate_journal_entry(
                            settlement_date, 1010, vat_account, abs(balance), np.nan,
                            f"Régularisation TVA compte {vat_account} Q{quarter}",
                            "", ref_document
                        ))
        return journal_entries

    def generate_misc_entries(self, year):
        """Generate miscellaneous entries with varied accounts"""
        journal_entries = []
        # Monthly entries (e.g., rent, cleaning, admin)
        for month in range(1, 13):
            date = self.get_random_date(year, month)
            # Rent (6000)
            rent = round(random.uniform(2000, 3000), 2)
            journal_entries.append(self.generate_journal_entry(
                date, 6000, 1010, rent, np.nan,
                f"Loyer mois {month}",
                "", f"RENT-{year}-{month:02d}"
            ))
            # Cleaning (6040, 50% of months)
            if random.random() < 0.5:
                cleaning = round(random.uniform(100, 300), 2)
                journal_entries.append(self.generate_journal_entry(
                    date, 6040, 1000, cleaning, np.nan,
                    f"Nettoyage mois {month}",
                    "", f"CLEAN-{year}-{month:02d}"
                ))
            # Admin fees (6500)
            admin = round(random.uniform(50, 200), 2)
            journal_entries.append(self.generate_journal_entry(
                date, 6500, 1010, admin, np.nan,
                f"Frais administratifs mois {month}",
                "", f"ADMIN-{year}-{month:02d}"
            ))
            # Internet (6510)
            internet = round(random.uniform(80, 120), 2)
            journal_entries.append(self.generate_journal_entry(
                date, 6510, 1010, internet, np.nan,
                f"Téléphone et internet mois {month}",
                "", f"INET-{year}-{month:02d}"
            ))
        # Quarterly entries
        for quarter in range(1, 5):
            date = datetime(year, (quarter - 1) * 3 + 3, 1)
            # Advertising (6600)
            advert = round(random.uniform(500, 1500), 2)
            journal_entries.append(self.generate_journal_entry(
                date, 6600, 1010, advert, np.nan,
                f"Publicité Q{quarter}",
                "", f"AD-{year}-Q{quarter}"
            ))
            # Client gifts (6643, ~10 times/year)
            if random.random() < 0.6:  # ~2-3 times per quarter
                gift = round(random.uniform(50, 200), 2)
                journal_entries.append(self.generate_journal_entry(
                    date, 6643, 1000, gift, np.nan,
                    f"Cadeaux clients Q{quarter}",
                    "", f"GIFT-{year}-Q{quarter}"
                ))
            # Amortization (6800)
            amort = round(random.uniform(300, 600), 2)
            journal_entries.append(self.generate_journal_entry(
                date, 6800, 1500, amort, np.nan,
                f"Amortissement Q{quarter}",
                "", f"AMORT-{year}-Q{quarter}"
            ))
        # Annual entries
        # Bank fees (6900, <18 times/year, ~12-15 times)
        for i in range(random.randint(12, 15)):
            date = self.get_random_date(year)
            fee = round(random.uniform(20, 100), 2)
            journal_entries.append(self.generate_journal_entry(
                date, 6900, 1010, fee, np.nan,
                f"Frais bancaires",
                "", f"FEE-{year}-{i:02d}"
            ))
        # Taxes (8900)
        date = datetime(year, 12, 31)
        tax = round(random.uniform(5000, 10000), 2)
        journal_entries.append(self.generate_journal_entry(
            date, 8900, 1010, tax, np.nan,
            f"Impôts directs {year}",
            "", f"TAX-{year}"
        ))
        # COVID-related entries
        if year == 2021:
            date = datetime(year, 3, 15)
            subsidy = round(random.uniform(10000, 20000), 2)
            journal_entries.append(self.generate_journal_entry(
                date, 1010, 8510, np.nan, subsidy,
                "Subvention COVID-19",
                "", "COV-2021-001"
            ))
        if year in [2021, 2022]:
            for i in range(3):  # ~3 safety expenses per year
                date = self.get_random_date(year)
                expense = round(random.uniform(100, 500), 2)
                journal_entries.append(self.generate_journal_entry(
                    date, 8200, 1000, expense, np.nan,
                    "Masques et désinfectants",
                    "", f"COV-{year}-{i:02d}"
                ))
        # Other charges (6700, ~20 times/year)
        for i in range(20):
            date = self.get_random_date(year)
            charge = round(random.uniform(50, 300), 2)
            journal_entries.append(self.generate_journal_entry(
                date, 6700, 1010, charge, np.nan,
                f"Autres charges",
                "", f"OTH-{year}-{i:02d}"
            ))
        return journal_entries

    def generate_accounting_data(self):
        """Generate all accounting data"""
        journal_entries = []
        factures_fournisseurs = []
        factures_clients = []
        numero_document = 1
        for year in range(self.params.start_year, self.params.end_year + 1):
            # Supplier invoices (~50/year)
            for i in range(1, 51):
                facture, entries = self.generate_supplier_invoice(year, numero_document)
                factures_fournisseurs.append(facture)
                journal_entries.extend(entries)
                numero_document += 1
            # Client invoices (~60/year)
            for i in range(1, 61):
                facture, entries = self.generate_client_invoice(year, numero_document)
                factures_clients.append(facture)
                journal_entries.extend(entries)
                numero_document += 1
            # Salaries
            journal_entries.extend(self.generate_salary_entries(year))
            # Miscellaneous entries
            journal_entries.extend(self.generate_misc_entries(year))
        journal_df = pd.DataFrame(journal_entries)
        # VAT settlements
        for year in range(self.params.start_year, self.params.end_year + 1):
            journal_entries.extend(self.generate_vat_settlement(journal_df, year))
        journal_df = pd.DataFrame(journal_entries)
        return journal_df, factures_fournisseurs, factures_clients

    def generate_financial_statements(self, journal_df):
        """Generate balance sheet and profit/loss statement with Solde column"""
        # EtatDeResultat (Profit and Loss)
        resultat_columns = ["Compte", "Intitule"]
        for year in range(self.params.start_year, self.params.end_year + 1):
            resultat_columns.append(f"MontantExercice{year - self.params.start_year + 1}")
        resultat_columns.append("Solde")  # Add Solde column
        resultat_rows = []
        for account_id, account_info in self.accounts.items():
            if account_info["category"] in ["Produit", "Charge"]:
                row = {"Compte": account_id, "Intitule": account_info["name"]}
                solde = 0.0
                for year in range(self.params.start_year, self.params.end_year + 1):
                    year_entries = journal_df[journal_df["Date"].str.contains(str(year))]
                    debit_sum = year_entries[year_entries["CompteDebit"] == account_id]["MontantDebit"].sum()
                    credit_sum = year_entries[year_entries["CompteCredit"] == account_id]["MontantCredit"].sum()
                    montant = credit_sum - debit_sum if account_info["category"] == "Produit" else debit_sum - credit_sum
                    # Products negative, charges positive
                    adjusted_montant = -montant if account_info["category"] == "Produit" else montant
                    row[f"MontantExercice{year - self.params.start_year + 1}"] = round(adjusted_montant, 2)
                    solde += adjusted_montant
                row["Solde"] = round(solde, 2)
                resultat_rows.append(row)
        resultat_df = pd.DataFrame(resultat_rows)

        # Calculate annual profit/loss for account 2979
        annual_results = {}
        for year in range(self.params.start_year, self.params.end_year + 1):
            year_idx = year - self.params.start_year + 1
            year_total = resultat_df[f"MontantExercice{year_idx}"].sum()
            annual_results[year] = round(year_total, 2)

        # Bilan (Balance Sheet)
        balance_columns = ["Compte", "Intitule"]
        for year in range(self.params.start_year, self.params.end_year + 1):
            balance_columns.append(f"SoldeExercice{year - self.params.start_year + 1}")
        balance_rows = []
        for account_id, account_info in self.accounts.items():
            if account_info["category"] in ["Actif", "Passif"] and account_id != 2979:
                row = {"Compte": account_id, "Intitule": account_info["name"]}
                for year in range(self.params.start_year, self.params.end_year + 1):
                    year_entries = journal_df[journal_df["Date"].str.contains(str(year))]
                    debit_sum = year_entries[year_entries["CompteDebit"] == account_id]["MontantDebit"].sum()
                    credit_sum = year_entries[year_entries["CompteCredit"] == account_id]["MontantCredit"].sum()
                    balance = debit_sum - credit_sum if account_info["category"] == "Actif" else credit_sum - debit_sum
                    row[f"SoldeExercice{year - self.params.start_year + 1}"] = round(balance, 2)
                balance_rows.append(row)
        # Add account 2979 (Result of the exercise)
        result_row = {"Compte": 2979, "Intitule": "Bénéfice/perte"}
        for year in range(self.params.start_year, self.params.end_year + 1):
            result_row[f"SoldeExercice{year - self.params.start_year + 1}"] = annual_results[year]
        balance_rows.append(result_row)
        balance_df = pd.DataFrame(balance_rows)
        return balance_df, resultat_df

    def generate_cotations_devises(self):
        """Generate currency exchange rates"""
        cotations = []
        for year in range(self.params.start_year, self.params.end_year + 1):
            for month in range(1, 13):
                date = datetime(year, month, 1)
                for code in ["EUR", "USD"]:
                    taux = round(random.uniform(0.85, 1.15), 2)  # Simplified rates
                    cotations.append({
                        "CodeMonnaie": code,
                        "DateCotation": date.strftime("%d.%m.%Y"),
                        "Taux": taux,
                        "Source": "Table de monnaies Abacus"
                    })
        return pd.DataFrame(cotations)

    def generate_all_data(self):
        """Generate all simulation data"""
        journal_df, factures_fournisseurs, factures_clients = self.generate_accounting_data()
        # PlanComptable
        plan_comptable_df = pd.DataFrame([
            {
                "Compte": k,
                "Intitule": v["name"],
                "Categorie": v["category"],
                "SousClasse": v["sous_classe"],
                "DetailCategorie": v["detail_categorie"],
                "TypeCompte": v["type_compte"]
            } for k, v in self.accounts.items()
        ])
        # CodesAnalytiques
        codes_analytiques_df = pd.DataFrame([
            {"Code": k, "Libelle": v["libelle"], "Type": v["type"]}
            for k, v in self.analytical_codes.items()
        ])
        # Fournisseurs
        fournisseurs_df = pd.DataFrame([
            {
                "IDFournisseur": k,
                "Nom": v["nom"],
                "Adresse": v["adresse"],
                "CodePostal": v["code_postal"],
                "Pays": v["pays"],
                "DeviseFacture": v["devise_facture"]
            } for k, v in self.suppliers.items()
        ])
        # FacturesFournisseurs
        factures_fournisseurs_df = pd.DataFrame(factures_fournisseurs)
        # Clients
        clients_df = pd.DataFrame([
            {
                "IDClient": k,
                "Nom": v["nom"],
                "Adresse": v["adresse"],
                "CodePostal": v["code_postal"],
                "Pays": v["pays"]
            } for k, v in self.clients.items()
        ])
        # FacturesClients
        factures_clients_df = pd.DataFrame(factures_clients)
        # Monnaies
        monnaies_df = pd.DataFrame([
            {"Code": k, "Nom": v["nom"]}
            for k, v in self.currencies.items()
        ])
        # CotationsDevises
        cotations_devises_df = self.generate_cotations_devises()
        # BalanceDesComptes and Bilan (same structure), EtatDeResultat
        balance_df, etat_resultat_df = self.generate_financial_statements(journal_df)
        return {
            "GrandLivre": journal_df,
            "PlanComptable": plan_comptable_df,
            "CodesAnalytiques": codes_analytiques_df,
            "Fournisseurs": fournisseurs_df,
            "FacturesFournisseurs": factures_fournisseurs_df,
            "Clients": clients_df,
            "FacturesClients": factures_clients_df,
            "Monnaies": monnaies_df,
            "CotationsDevises": cotations_devises_df,
            "BalanceDesComptes": balance_df,
            "Bilan": balance_df,
            "EtatDeResultat": etat_resultat_df
        }

# Excel utility
class ExcelGenerator:
    def __init__(self, params):
        self.params = params

    def create_excel_file(self, data, file_name):
        """Create Excel file in the current directory"""
        file_path = file_name  # Save directly in current directory
        wb = Workbook()
        if wb.sheetnames:
            wb.remove(wb.active)
        for table_name, df in data.items():
            ws = wb.create_sheet(title=table_name)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            for col_idx, column_cells in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for cell in column_cells:
                    try:
                        value = str(cell.value) if cell.value is not None else ""
                        max_length = max(max_length, len(value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            if not df.empty:
                max_row = ws.max_row
                max_col = ws.max_column
                last_col_letter = get_column_letter(max_col)
                ref_range = f"A1:{last_col_letter}{max_row}"
                table = Table(displayName=f"Table_{table_name.replace(' ', '_')}", ref=ref_range)
                style = TableStyleInfo(
                    name="TableStyleMedium11",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)
        wb.save(file_path)
        logging.info(f"Excel file created: {file_path}")
        return file_path

    def export_all_data(self, data):
        """Export all data to Excel"""
        file_name = "BookstoreAccountingData.xlsx"
        file_path = self.create_excel_file(data, file_name)
        return {file_name: file_path}

# Main program
def main():
    params = SimulationParams(start_year=2021, end_year=2025)
    logging.info(f"Starting bookstore accounting simulation {params.start_year}-{params.end_year}")
    generator = AccountingDataGenerator(params)
    data = generator.generate_all_data()
    excel_generator = ExcelGenerator(params)
    exported_files = excel_generator.export_all_data(data)
    logging.info(f"Files exported to current directory")
    logging.info(f"Total journal entries: {len(data['GrandLivre'])}")
    logging.info(f"Number of accounts: {len(data['PlanComptable'])}")

if __name__ == "__main__":
    main()
